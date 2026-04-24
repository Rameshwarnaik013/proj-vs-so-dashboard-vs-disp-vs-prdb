import openpyxl
import json
import os
import time

def process_excel_data(file_path, progress_callback=None):
    """
    Parses the 76MB Excel file using openpyxl streaming (read_only) mode.
    This is significantly faster and more RAM-efficient than pandas.read_excel.
    """
    def log(msg):
        print(msg)
        if progress_callback:
            progress_callback(msg)

    log(f"[*] Starting high-speed streaming read of {os.path.basename(file_path)}...")
    start_time = time.time()

    MONTHS = ['Apr-25','May-25','Jun-25','Jul-25','Aug-25','Sep-25','Oct-25','Nov-25','Dec-25','Jan-26','Feb-26','Mar-26']
    DIMENSIONS = [
        'Item Parent', 'Customer', 'Customer Group', 'Origin', 
        'New Mis Item Group', 'Item Type(KVI/VALUE ADDED)', 
        'Packaging Type ', 'Packaging Method', 'Sales Order Created By',
        'Item Name'
    ]

    # Structure to hold the aggregates
    dim_data = { d: {} for d in DIMENSIONS }

    try:
        # Load workbook in read-only streaming mode
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        all_sheets = wb.sheetnames
        
        proj_sheet_name = next((s for s in all_sheets if 'proj' in s.lower()), None)
        so_sheet_name = next((s for s in all_sheets if 'so' in s.lower() and 'dispatch' in s.lower()), None)
        prdn_sheet_name = next((s for s in all_sheets if 'prdn' in s.lower() or 'prod' in s.lower()), None)

        if not all([proj_sheet_name, so_sheet_name, prdn_sheet_name]):
            raise ValueError(f"Missing required sheets. Found: {all_sheets}")

        def parse_sheet(sheet_name, extra_metrics):
            log(f"Reading {sheet_name}...")
            ws = wb[sheet_name]
            
            # 1. Find Header Row and Map Columns
            header_row = None
            h_row_idx = -1
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10)):
                values = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
                if 'MMM-YY' in values or 'MMM - YY' in values:
                    header_row = values
                    h_row_idx = i + 1
                    break
            
            if not header_row:
                log(f"[!] Warning: Header not found in {sheet_name}.")
                return

            date_idx = -1
            date_col = 'MMM-YY' if 'MMM-YY' in header_row else 'MMM - YY'
            date_idx = header_row.index(date_col)
            
            dim_map = {}
            for d in DIMENSIONS:
                actual = next((h for h in header_row if h.strip() == d.strip()), None)
                if actual: dim_map[d] = header_row.index(actual)

            metric_map = {}
            for m_source, m_target in extra_metrics.items():
                if m_source in header_row:
                    metric_map[header_row.index(m_source)] = m_target

            # 2. Iterate Rows
            for i, row in enumerate(ws.iter_rows(min_row=h_row_idx + 1)):
                if (i + 1) % 10000 == 0:
                    log(f"    -> {sheet_name}: {i + 1} rows...")
                
                # Check for empty data
                if len(row) <= date_idx or not row[date_idx].value:
                    continue

                # Normalize Date
                raw_date = row[date_idx].value
                try:
                    if hasattr(raw_date, 'strftime'):
                        date_val = raw_date.strftime('%b-%y')
                    else:
                        date_val = str(raw_date).strip()
                except: continue

                if date_val not in MONTHS: continue

                # Process Metrics for each dimension
                for dim, d_idx in dim_map.items():
                    if d_idx >= len(row): continue
                    val_name = str(row[d_idx].value).strip() if row[d_idx].value is not None else "Unknown"
                    if not val_name or val_name == 'None': val_name = "Unknown"
                    if val_name == 'Jar Sealing - WAD Machine': val_name = 'Jar Sealing - WAD Machine/Table'

                    if val_name not in dim_data[dim]:
                        dim_data[dim][val_name] = { m: { 'proj':0,'so':0,'disp':0,'pend':0,'clsd':0,'prdn':0, 'proj_qty':0,'so_qty':0,'disp_qty':0,'pend_qty':0,'clsd_qty':0,'prdn_qty':0 } for m in MONTHS }
                    
                    target_month = dim_data[dim][val_name][date_val]
                    for m_idx, m_target in metric_map.items():
                        if m_idx < len(row):
                            val = row[m_idx].value
                            try:
                                target_month[m_target] += float(val) if val is not None else 0
                            except: pass

        # Load each sheet
        # Kg-based metrics
        parse_sheet(proj_sheet_name, {'Stock Qty In Kg':'proj', 'Projection Units':'proj_qty'})
        parse_sheet(so_sheet_name, {'Stock Qty In Kg':'so', 'Qty':'so_qty', 'Delivered KGs':'disp', 'Delivered Qty':'disp_qty', 'Pending KGs':'pend', 'Closed KGs':'clsd'})
        parse_sheet(prdn_sheet_name, {'Stock Qty In Kg':'prdn', 'Qty':'prdn_qty'})

        # Derive pending_qty and closed_qty from so_qty and disp_qty
        for dim in DIMENSIONS:
            for val_name, months_data in dim_data[dim].items():
                for m in MONTHS:
                    md = months_data[m]
                    # Pending Qty = SO Qty - Dispatched Qty (clamped to 0)
                    md['pend_qty'] = max(0, md['so_qty'] - md['disp_qty'])
                    # Closed Qty = SO Qty - Dispatched Qty - Pending Qty (or derive similarly to KGs ratio)
                    # Use same ratio as KGs: if so > 0, clsd_qty = clsd / so * so_qty
                    if md['so'] > 0 and md['clsd'] > 0:
                        md['clsd_qty'] = round(md['clsd'] / md['so'] * md['so_qty'], 2)
                    else:
                        md['clsd_qty'] = 0

        wb.close()

        # --- Aggregation logic ---
        log("[*] Finalizing calculations...")
        
        ALL_KEYS = ['proj', 'so', 'disp', 'pend', 'clsd', 'prdn', 'proj_qty', 'so_qty', 'disp_qty', 'pend_qty', 'clsd_qty', 'prdn_qty']
        monthly = { 'months': MONTHS }
        for k in ALL_KEYS:
            monthly[k] = []
        
        for m in MONTHS:
            dim0 = DIMENSIONS[0]
            for k in ALL_KEYS:
                val = sum(v[m][k] for v in dim_data[dim0].values())
                monthly[k].append(round(val, 2))

        kpis = { k: sum(monthly[k]) for k in ALL_KEYS }

        tables = {}
        for d in DIMENSIONS:
            rows = []
            for name, months_data in dim_data[d].items():
                row = {'name': name}
                for k in ALL_KEYS:
                    row[k] = sum(mdata[k] for mdata in months_data.values())
                row['so_proj_pct'] = round((row['so']/row['proj']*100), 1) if row['proj'] > 0 else 0
                row['disp_so_pct'] = round((row['disp']/row['so']*100), 1) if row['so'] > 0 else 0
                row['so_proj_pct_qty'] = round((row['so_qty']/row['proj_qty']*100), 1) if row['proj_qty'] > 0 else 0
                row['disp_so_pct_qty'] = round((row['disp_qty']/row['so_qty']*100), 1) if row['so_qty'] > 0 else 0
                rows.append(row)
            tables[d] = sorted(rows, key=lambda x: x['proj'], reverse=True)

        filters = { d: sorted(list(dim_data[d].keys())) for d in DIMENSIONS }
        
        QTR_ORDER = ['Q1 (Apr-Jun)', 'Q2 (Jul-Sep)', 'Q3 (Oct-Dec)', 'Q4 (Jan-Mar)']
        QMAP_M = {
            'Apr-25':0, 'May-25':0, 'Jun-25':0,
            'Jul-25':1, 'Aug-25':1, 'Sep-25':1,
            'Oct-25':2, 'Nov-25':2, 'Dec-25':2,
            'Jan-26':3, 'Feb-26':3, 'Mar-26':3
        }
        quarterly = { 'quarters': QTR_ORDER }
        for k in ALL_KEYS:
            quarterly[k] = [0]*4
        for m in MONTHS:
            idx = QMAP_M[m]
            m_idx = MONTHS.index(m)
            for k in ALL_KEYS:
                quarterly[k][idx] += monthly[k][m_idx]

        final_data = {
            'kpis': kpis,
            'monthly': monthly,
            'quarterly': quarterly,
            'tables': tables,
            'filters': filters,
            'dim_data': dim_data
        }

        duration = time.time() - start_time
        log(f"[*] Done in {duration:.2f}s.")
        return final_data

    except Exception as e:
        log(f"Error processing Excel data: {e}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    FILE = r"c:\Users\Admin\OneDrive - https farmley.com\Desktop\FINAL_01\Projection vs so vs disp vs prdn dashboard.xlsx"
    result = process_excel_data(FILE)
    if result:
        html_path = r"c:\Users\Admin\OneDrive - https farmley.com\Desktop\FINAL_01\index.html"
        with open(html_path, 'r', encoding='utf-8') as f:
            html = f.read()
        # Find and replace the EMBEDDED_DATA line using string operations (regex fails on backslash escapes in JSON)
        marker = 'const EMBEDDED_DATA = '
        start_idx = html.index(marker)
        # Find the matching closing ";", scanning for the end of the JSON object
        brace_count = 0
        i = html.index('{', start_idx)
        while i < len(html):
            if html[i] == '{': brace_count += 1
            elif html[i] == '}': brace_count -= 1
            if brace_count == 0:
                end_idx = html.index(';', i) + 1
                break
            i += 1
        json_str = json.dumps(result, separators=(',', ': '))
        new_line = f'const EMBEDDED_DATA = {json_str};'
        html = html[:start_idx] + new_line + html[end_idx:]
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(html)
        print(f"[*] Patched index.html with fresh data ({len(json_str)//1024} KB)")
