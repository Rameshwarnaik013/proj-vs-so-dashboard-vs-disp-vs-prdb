import openpyxl
import json
import os
import time

def process_excel_data(file_path, progress_callback=None):
    """
    Parses the 76MB Excel file using openpyxl streaming (read_only) mode.
    This is significantly faster and more RAM-efficient than pandas.read_excel.
    """
    def log(msg):
        print(msg)
        if progress_callback:
            progress_callback(msg)

    log(f"[*] Starting high-speed streaming read of {os.path.basename(file_path)}...")
    start_time = time.time()

    MONTHS = ['Apr-25','May-25','Jun-25','Jul-25','Aug-25','Sep-25','Oct-25','Nov-25','Dec-25','Jan-26','Feb-26','Mar-26','Apr-26','May-26']
    DIMENSIONS = [
        'Item Parent', 'Customer', 'Customer Group', 'Origin',
        'New Mis Item Group', 'Item Type(KVI/VALUE ADDED)',
        'Packaging Type ', 'Packaging Method', 'Sales Order Created By',
        'Item Name', 'Origin + Item Name'
    ]

    # Fixed vocabulary for deriving Packaging Type from Item Name text (Projection tab, then
    # propagated to SO/dispatch and Prdn by Item Code). Order doesn't matter here - matching
    # is always tried longest-normalized-first so "Composite Jar"/"Tin Jar" beat bare "Jar".
    PACKAGING_TYPE_KEYWORDS = [
        'Standee Pouch', 'Monocarton', 'Jar', 'Tin Jar', 'Pillow pouch',
        'Gifting Tray', 'Potli-Gifting', 'Gifting Box', 'Center Seal Pouch', 'Composite Jar'
    ]

    # Business-rule table: (Origin, Item Type(KVI/VALUE ADDED), Packaging Type) -> Packaging Method.
    # A couple of combos in the source table were ambiguous (same combo, different Method) and
    # were resolved as directed: Indore/Value Added/Standee Pouch -> PFS only (the stray "Table"
    # entry is dropped); Indore/Value Added/Jar is resolved at match-time via _resolve_method_for_type
    # (Conti Mix when Item Name mentions Panchmeva/Trail Mix, else Jar Sealing - WAD Machine/Table).
    # Indore/KVI/Jar had a similar stray "Table" entry with no guidance given - defaulted to
    # "Jar Sealing - WAD Machine/Table" for consistency with every other Jar combo; flag if wrong.
    ORIGIN_TYPE_METHOD = {
        ('indore', 'value added', 'Standee Pouch'): 'PFS',
        ('indore', 'kvi', 'Standee Pouch'): 'Table',
        ('indore', 'value added', 'Monocarton'): 'Table',
        ('indore', 'value added', 'Tin Jar'): 'Table',
        ('indore', 'kvi', 'Tin Jar'): 'Table',
        ('indore', 'kvi', 'Generic Pouch'): 'Discontinued',
        ('indore', 'kvi', 'Jar'): 'Jar Sealing - WAD Machine/Table',
        ('indore', 'value added', 'Pillow pouch'): 'FFS',
        ('indore', 'kvi', 'Pillow pouch'): 'FFS',
        ('indore', 'value added', 'Gifting Tray'): 'Table',
        ('indore', 'value added', 'Potli-Gifting'): 'Table',
        ('indore', 'value added', 'Gifting Box'): 'Table',
        ('indore', 'kvi', 'Gifting Box'): 'Table',
        ('purnea', 'value added', 'Jar'): 'Jar Sealing - WAD Machine/Table',
        ('purnea', 'kvi', 'Center Seal Pouch'): 'Table',
        ('purnea', 'value added', 'Composite Jar'): 'Seaming',
        ('purnea', 'value added', 'Pillow pouch'): 'FFS',
        ('purnea', 'value added', 'Center Seal Pouch'): 'Table',
        ('purnea', 'value added', 'Standee Pouch'): 'PFS',
        ('ud foods', 'value added', 'Pillow pouch'): 'FFS',
        ('udupi', 'kvi', 'Standee Pouch'): 'Table',
        ('udupi', 'kvi', 'Generic Pouch'): 'Discontinued',
        ('functional foods', 'value added', 'Pillow pouch'): 'FFS',
    }

    # SO/dispatch rows matching any of these are excluded from every dashboard total (not just
    # packaging fields) - treated as if they never existed in the source data.
    SO_DROP_WORKFLOW_STATE = {'internal transfer', 'on hold', 'rejected', 'cancelled'}
    SO_DROP_SO_STATUS = {'cancelled', 'on hold', 'internal transfer', 'rejected'}

    # Structure to hold the aggregates
    dim_data = { d: {} for d in DIMENSIONS }

    try:
        # Load workbook in read-only streaming mode
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        all_sheets = wb.sheetnames

        proj_sheet_name = next((s for s in all_sheets if 'proj' in s.lower()), None)
        so_sheet_name = next((s for s in all_sheets if 'so' in s.lower() and 'dispatch' in s.lower()), None)
        prdn_sheet_name = next((s for s in all_sheets if 'prdn' in s.lower() or 'prod' in s.lower()), None)
        lookup_sheet_name = next((s for s in all_sheets if 'lookup' in s.lower()), None)

        if not all([proj_sheet_name, so_sheet_name, prdn_sheet_name]):
            raise ValueError(f"Missing required sheets. Found: {all_sheets}")

        def _normalize(s):
            # Treat '_', '-' and repeated whitespace as equivalent when matching item codes/names
            if s is None:
                return ""
            s = str(s).strip().lower().replace('_', ' ').replace('-', ' ')
            return ' '.join(s.split())

        def _resolve_method_for_type(origin_val, item_type_val, packaging_type, name_norm):
            origin_n = _normalize(origin_val)
            type_n = _normalize(item_type_val)
            if origin_n == 'indore' and type_n == 'value added' and packaging_type == 'Jar':
                if 'panchmeva' in name_norm or 'trail mix' in name_norm:
                    return 'Conti Mix'
                return 'Jar Sealing - WAD Machine/Table'
            return ORIGIN_TYPE_METHOD.get((origin_n, type_n, packaging_type))

        # Reference map from the 'lookup' sheet: Item Code -> Packaging Type / Packaging Method.
        # Used as a fallback layer (after the Projection VLOOKUP) for codes not covered there.
        # Also builds a Type -> most common Method fallback for when neither Item Code nor the
        # (Origin, Item Type, Packaging Type) business rule table can resolve a Method.
        code_to_type = {}
        code_to_method = {}
        type_method_counts = {}
        if lookup_sheet_name:
            log(f"Reading {lookup_sheet_name} (packaging reference)...")
            ws_lu = wb[lookup_sheet_name]
            for lu_row in ws_lu.iter_rows(min_row=2, max_row=5000, values_only=True):
                if not lu_row or lu_row[0] is None:
                    continue
                code_norm = _normalize(lu_row[0])
                ptype = str(lu_row[1]).strip() if len(lu_row) > 1 and lu_row[1] is not None else ""
                pmethod = str(lu_row[2]).strip() if len(lu_row) > 2 and lu_row[2] is not None else ""
                if ptype:
                    code_to_type[code_norm] = ptype
                if pmethod:
                    code_to_method[code_norm] = pmethod
                if ptype and pmethod:
                    type_method_counts.setdefault(ptype, {})
                    type_method_counts[ptype][pmethod] = type_method_counts[ptype].get(pmethod, 0) + 1

        type_to_method_mode = {
            t: max(counts.items(), key=lambda kv: kv[1])[0]
            for t, counts in type_method_counts.items()
        }
        # Longest keyword first so "Composite Jar" / "Tin Jar" match before bare "Jar"
        type_keywords_norm = sorted(
            ((_normalize(t), t) for t in PACKAGING_TYPE_KEYWORDS),
            key=lambda p: -len(p[0])
        )

        # Item Code -> (Packaging Type, Packaging Method) resolved on the Projection tab.
        # Populated while parsing Projection (parsed first), then used as the primary VLOOKUP
        # source when parsing SO/dispatch and Prdn, so all three tabs agree by Item Code.
        projection_code_map = {}

        def parse_sheet(sheet_name, extra_metrics):
            log(f"Reading {sheet_name}...")
            ws = wb[sheet_name]
            is_projection = (sheet_name == proj_sheet_name)
            is_so = (sheet_name == so_sheet_name)

            # 1. Find Header Row and Map Columns
            header_row = None
            h_row_idx = -1
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10)):
                values = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
                if 'MMM-YY' in values or 'MMM - YY' in values:
                    header_row = values
                    h_row_idx = i + 1
                    break

            if not header_row:
                log(f"[!] Warning: Header not found in {sheet_name}.")
                return

            date_idx = -1
            date_col = 'MMM-YY' if 'MMM-YY' in header_row else 'MMM - YY'
            date_idx = header_row.index(date_col)

            dim_map = {}
            for d in DIMENSIONS:
                actual = next((h for h in header_row if h.strip() == d.strip()), None)
                if actual: dim_map[d] = header_row.index(actual)

            metric_map = {}
            for m_source, m_target in extra_metrics.items():
                if m_source in header_row:
                    metric_map[header_row.index(m_source)] = m_target

            def find_col(name):
                return next((header_row.index(h) for h in header_row if h.strip() == name), -1)

            origin_idx = find_col('Origin')
            item_idx = find_col('Item Name')
            itemcode_idx = find_col('Item Code')
            item_group_idx = find_col('Item Group')
            item_type_idx = find_col('Item Type(KVI/VALUE ADDED)')
            type_col_idx = dim_map.get('Packaging Type ', -1)
            method_col_idx = dim_map.get('Packaging Method', -1)

            # SO/dispatch row-filter columns (only used when is_so)
            created_by_idx = find_col('Sales Order Created By')
            workflow_state_idx = find_col('Workflow State')
            purpose_idx = find_col('Purpose')
            so_status_idx = find_col('Sales Order Status')
            customer_idx = find_col('Customer')
            item_type_bare_idx = find_col('Item Type')
            returned_qty_idx = find_col('Returned Qty')
            filtered_count = 0

            def cellstr(idx, row):
                return str(row[idx].value).strip() if idx >= 0 and idx < len(row) and row[idx].value is not None else ""

            # 2. Iterate Rows
            for i, row in enumerate(ws.iter_rows(min_row=h_row_idx + 1)):
                if (i + 1) % 10000 == 0:
                    log(f"    -> {sheet_name}: {i + 1} rows...")

                # Check for empty data
                if len(row) <= date_idx or not row[date_idx].value:
                    continue

                # Normalize Date
                raw_date = row[date_idx].value
                try:
                    if hasattr(raw_date, 'strftime'):
                        date_val = raw_date.strftime('%b-%y')
                    else:
                        date_val = str(raw_date).strip()
                except: continue

                if date_val not in MONTHS: continue

                # SO/dispatch: drop non-genuine sales rows before they count toward anything
                # (admin test orders, cancelled/on-hold/rejected/internal-transfer, documentation
                # purpose, sample orders, bulk item type, over-returned qty).
                if is_so:
                    if cellstr(created_by_idx, row).lower() == 'administrator':
                        filtered_count += 1; continue
                    if cellstr(workflow_state_idx, row).lower() in SO_DROP_WORKFLOW_STATE:
                        filtered_count += 1; continue
                    if cellstr(purpose_idx, row).lower() == 'documentation':
                        filtered_count += 1; continue
                    if cellstr(so_status_idx, row).lower() in SO_DROP_SO_STATUS:
                        filtered_count += 1; continue
                    if cellstr(customer_idx, row).lower().startswith('sample order'):
                        filtered_count += 1; continue
                    if cellstr(item_type_bare_idx, row).lower() == 'bulk':
                        filtered_count += 1; continue
                    if returned_qty_idx >= 0 and returned_qty_idx < len(row):
                        rq_raw = row[returned_qty_idx].value
                        try:
                            if rq_raw is not None and float(rq_raw) > 1:
                                filtered_count += 1; continue
                        except (TypeError, ValueError):
                            pass

                item_code_val = cellstr(itemcode_idx, row)
                item_name_val = cellstr(item_idx, row)
                code_norm = _normalize(item_code_val)
                name_norm = _normalize(item_name_val)

                # Resolve Packaging Type/Method once per row.
                resolved_type = None
                resolved_method = None

                if is_projection:
                    # Projection is the master: Type from Item Name keywords, Method from the
                    # (Origin, Item Type, Packaging Type) business rule table.
                    if name_norm:
                        for kw_norm, kw_orig in type_keywords_norm:
                            if kw_norm and kw_norm in name_norm:
                                resolved_type = kw_orig
                                break
                    if not resolved_type and type_col_idx >= 0:
                        raw_t = cellstr(type_col_idx, row)
                        if raw_t: resolved_type = raw_t

                    if resolved_type:
                        origin_val = cellstr(origin_idx, row)
                        item_type_val = cellstr(item_type_idx, row)
                        resolved_method = _resolve_method_for_type(origin_val, item_type_val, resolved_type, name_norm)
                    if not resolved_method and resolved_type:
                        resolved_method = type_to_method_mode.get(resolved_type)
                    if not resolved_method and method_col_idx >= 0:
                        raw_m = cellstr(method_col_idx, row)
                        if raw_m: resolved_method = raw_m

                    if code_norm:
                        projection_code_map[code_norm] = (resolved_type or "Unknown", resolved_method or "Unknown")
                else:
                    # SO/dispatch and Prdn: VLOOKUP from Projection's resolved values first (by
                    # Item Code), then the row's own cell, then the lookup sheet, then keyword match.
                    proj_match = projection_code_map.get(code_norm) if code_norm else None
                    if proj_match:
                        resolved_type, resolved_method = proj_match

                    if not resolved_type and type_col_idx >= 0:
                        raw_t = cellstr(type_col_idx, row)
                        if raw_t: resolved_type = raw_t
                    if not resolved_method and method_col_idx >= 0:
                        raw_m = cellstr(method_col_idx, row)
                        if raw_m: resolved_method = raw_m

                    if not resolved_type:
                        resolved_type = code_to_type.get(code_norm)
                    if not resolved_method:
                        resolved_method = code_to_method.get(code_norm)

                    if not resolved_type and name_norm:
                        for kw_norm, kw_orig in type_keywords_norm:
                            if kw_norm and kw_norm in name_norm:
                                resolved_type = kw_orig
                                break

                    if not resolved_method and resolved_type:
                        resolved_method = type_to_method_mode.get(resolved_type)

                # Process Metrics for each dimension
                for dim in DIMENSIONS:
                    val_name = "Unknown"
                    if dim == 'Origin + Item Name':
                        o_val = cellstr(origin_idx, row) or "Unknown"
                        i_val = item_name_val or "Unknown"
                        val_name = f"[{o_val}] {i_val}"
                    elif dim == 'Packaging Type ':
                        val_name = resolved_type or "Unknown"
                    elif dim == 'Packaging Method':
                        val_name = resolved_method or "Unknown"
                    elif dim == 'New Mis Item Group' and is_projection:
                        # Projection tab: Item Group is treated as New Mis Item Group
                        val_name = cellstr(item_group_idx, row) or "Unknown"
                    else:
                        d_idx = dim_map.get(dim, -1)
                        if d_idx >= 0 and d_idx < len(row):
                            val_name = str(row[d_idx].value).strip() if row[d_idx].value is not None else "Unknown"

                    if not val_name or val_name == 'None': val_name = "Unknown"
                    if val_name == 'Jar Sealing - WAD Machine': val_name = 'Jar Sealing - WAD Machine/Table'

                    if val_name not in dim_data[dim]:
                        dim_data[dim][val_name] = { m: { 'proj':0,'so':0,'disp':0,'pend':0,'clsd':0,'prdn':0, 'proj_qty':0,'so_qty':0,'disp_qty':0,'pend_qty':0,'clsd_qty':0,'prdn_qty':0, 'so_amt':0 } for m in MONTHS }

                    target_month = dim_data[dim][val_name][date_val]
                    for m_idx, m_target in metric_map.items():
                        if m_idx < len(row):
                            val = row[m_idx].value
                            try:
                                target_month[m_target] += float(val) if val is not None else 0
                            except: pass

            if is_projection:
                log(f"    -> Projection packaging map built for {len(projection_code_map)} item codes")
            if is_so and filtered_count:
                log(f"    -> {sheet_name}: filtered out {filtered_count} non-genuine rows (admin/cancelled/on-hold/rejected/internal-transfer/documentation/sample/bulk/returns)")

        # Load each sheet (Projection first - it builds the packaging map SO/Prdn VLOOKUP against)
        # Kg-based metrics
        parse_sheet(proj_sheet_name, {'Stock Qty In Kg':'proj', 'Projection Units':'proj_qty'})
        parse_sheet(so_sheet_name, {'Stock Qty In Kg':'so', 'Qty':'so_qty', 'Delivered KGs':'disp', 'Delivered Qty':'disp_qty', 'Pending KGs':'pend', 'Closed KGs':'clsd', 'Amount':'so_amt'})
        parse_sheet(prdn_sheet_name, {'Stock Qty In Kg':'prdn', 'Qty':'prdn_qty'})

        # Derive pending_qty and closed_qty from so_qty and disp_qty
        for dim in DIMENSIONS:
            for val_name, months_data in dim_data[dim].items():
                for m in MONTHS:
                    md = months_data[m]
                    # Pending Qty = SO Qty - Dispatched Qty (clamped to 0)
                    md['pend_qty'] = max(0, md['so_qty'] - md['disp_qty'])
                    # Closed Qty = SO Qty - Dispatched Qty - Pending Qty (or derive similarly to KGs ratio)
                    # Use same ratio as KGs: if so > 0, clsd_qty = clsd / so * so_qty
                    if md['so'] > 0 and md['clsd'] > 0:
                        md['clsd_qty'] = round(md['clsd'] / md['so'] * md['so_qty'], 2)
                    else:
                        md['clsd_qty'] = 0

        wb.close()

        # --- Aggregation logic ---
        log("[*] Finalizing calculations...")

        ALL_KEYS = ['proj', 'so', 'disp', 'pend', 'clsd', 'prdn', 'proj_qty', 'so_qty', 'disp_qty', 'pend_qty', 'clsd_qty', 'prdn_qty', 'so_amt']
        monthly = { 'months': MONTHS }
        for k in ALL_KEYS:
            monthly[k] = []

        for m in MONTHS:
            dim0 = DIMENSIONS[0]
            for k in ALL_KEYS:
                val = sum(v[m][k] for v in dim_data[dim0].values())
                monthly[k].append(round(val, 2))

        kpis = { k: sum(monthly[k]) for k in ALL_KEYS }

        tables = {}
        for d in DIMENSIONS:
            rows = []
            for name, months_data in dim_data[d].items():
                row = {'name': name}
                for k in ALL_KEYS:
                    row[k] = sum(mdata[k] for mdata in months_data.values())
                row['so_proj_pct'] = round((row['so']/row['proj']*100), 1) if row['proj'] > 0 else 0
                row['disp_so_pct'] = round((row['disp']/row['so']*100), 1) if row['so'] > 0 else 0
                row['so_proj_pct_qty'] = round((row['so_qty']/row['proj_qty']*100), 1) if row['proj_qty'] > 0 else 0
                row['disp_so_pct_qty'] = round((row['disp_qty']/row['so_qty']*100), 1) if row['so_qty'] > 0 else 0
                rows.append(row)
            tables[d] = sorted(rows, key=lambda x: x['proj'], reverse=True)

        filters = { d: sorted(list(dim_data[d].keys())) for d in DIMENSIONS }

        QTR_ORDER = ['Q1 (Apr-Jun)', 'Q2 (Jul-Sep)', 'Q3 (Oct-Dec)', 'Q4 (Jan-Mar)', 'Q1 FY27 (Apr-Jun)']
        QMAP_M = {
            'Apr-25':0, 'May-25':0, 'Jun-25':0,
            'Jul-25':1, 'Aug-25':1, 'Sep-25':1,
            'Oct-25':2, 'Nov-25':2, 'Dec-25':2,
            'Jan-26':3, 'Feb-26':3, 'Mar-26':3,
            'Apr-26':4, 'May-26':4
        }
        quarterly = { 'quarters': QTR_ORDER }
        for k in ALL_KEYS:
            quarterly[k] = [0]*5
        for m in MONTHS:
            idx = QMAP_M[m]
            m_idx = MONTHS.index(m)
            for k in ALL_KEYS:
                quarterly[k][idx] += monthly[k][m_idx]

        final_data = {
            'kpis': kpis,
            'monthly': monthly,
            'quarterly': quarterly,
            'tables': tables,
            'filters': filters,
            'dim_data': dim_data
        }

        duration = time.time() - start_time
        log(f"[*] Done in {duration:.2f}s.")
        return final_data

    except Exception as e:
        log(f"Error processing Excel data: {e}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    FILE = r"c:\Users\Admin\OneDrive - https farmley.com\Desktop\FINAL_01\Projection vs so vs disp vs prdn dashboard (1).xlsx"
    result = process_excel_data(FILE)
    if result:
        html_path = r"c:\Users\Admin\OneDrive - https farmley.com\Desktop\FINAL_01\index.html"
        with open(html_path, 'r', encoding='utf-8') as f:
            html = f.read()
        # Find and replace the EMBEDDED_DATA line using string operations (regex fails on backslash escapes in JSON)
        marker = 'const EMBEDDED_DATA = '
        start_idx = html.index(marker)
        # Find the matching closing ";", scanning for the end of the JSON object
        brace_count = 0
        i = html.index('{', start_idx)
        while i < len(html):
            if html[i] == '{': brace_count += 1
            elif html[i] == '}': brace_count -= 1
            if brace_count == 0:
                end_idx = html.index(';', i) + 1
                break
            i += 1
        json_str = json.dumps(result, separators=(',', ': '))
        new_line = f'const EMBEDDED_DATA = {json_str};'
        html = html[:start_idx] + new_line + html[end_idx:]
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(html)
        print(f"[*] Patched index.html with fresh data ({len(json_str)//1024} KB)")
