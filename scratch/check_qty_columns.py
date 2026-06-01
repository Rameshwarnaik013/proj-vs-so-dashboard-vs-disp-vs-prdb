import openpyxl
import os

FILE_PATH = r"c:\Users\Admin\OneDrive - https farmley.com\Desktop\FINAL_01\Projection vs so vs disp vs prdn dashboard.xlsx"

wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)

# Check SO sheet - look for Pending Qty, Closed Qty columns
ws = wb['SO and dispatch ']
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=1)):
    values = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
    # Print columns that have "pend" or "clos" or "deliver" or "qty"
    for j, v in enumerate(values):
        if any(k in v.lower() for k in ['pend', 'clos', 'deliver', 'qty', 'unit', 'proj']):
            print(f"SO sheet: [{j}] '{v}'")

# Check Projection sheet for Qty-like columns  
ws2 = wb['Projection']
for i, row in enumerate(ws2.iter_rows(min_row=1, max_row=1)):
    values = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
    for j, v in enumerate(values):
        if any(k in v.lower() for k in ['qty', 'unit', 'proj', 'plan']):
            print(f"Proj sheet: [{j}] '{v}'")

# Check Prdn sheet for Qty-like columns
ws3 = wb['Prdn']
for i, row in enumerate(ws3.iter_rows(min_row=1, max_row=1)):
    values = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
    for j, v in enumerate(values):
        if any(k in v.lower() for k in ['qty', 'unit', 'produc', 'count']):
            print(f"Prdn sheet: [{j}] '{v}'")

# Also print a few sample values to understand the data
print("\n--- Sample data from SO sheet (first 3 data rows) ---")
ws_so = wb['SO and dispatch ']
header_found = False
for i, row in enumerate(ws_so.iter_rows(min_row=1, max_row=5)):
    values = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
    if not header_found and ('MMM-YY' in values or 'MMM - YY' in values):
        header_found = True
        continue
    if header_found:
        # Print Stock Qty In Kg (18), Qty (19), Delivered Qty (28)
        print(f"  Row {i+1}: StockQtyKg={values[18] if len(values)>18 else 'N/A'}, Qty={values[19] if len(values)>19 else 'N/A'}, DeliveredQty={values[28] if len(values)>28 else 'N/A'}")

# Check Projection sheet sample
print("\n--- Sample data from Projection sheet (first 3 data rows) ---")
ws_proj = wb['Projection']
header_found = False
for i, row in enumerate(ws_proj.iter_rows(min_row=1, max_row=5)):
    values = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
    if not header_found and ('MMM-YY' in values or 'MMM - YY' in values):
        header_found = True
        continue
    if header_found:
        # Print Stock Qty In Kg (11), Projection Units (8)
        print(f"  Row {i+1}: StockQtyKg={values[11] if len(values)>11 else 'N/A'}, ProjUnits={values[8] if len(values)>8 else 'N/A'}")

wb.close()
